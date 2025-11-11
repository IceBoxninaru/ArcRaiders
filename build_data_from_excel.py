#!/usr/bin/env python3
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

def read_sheet(ws):
    rows = []
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        rows.append({h:v for h,v in zip(headers, r)})
    return rows

ap = argparse.ArgumentParser()
ap.add_argument("--excel", required=True)
ap.add_argument("--out", required=True)
a = ap.parse_args()

wb = load_workbook(a.excel, data_only=True)
g = lambda n: wb[n] if n in wb.sheetnames else None

data = {
  "weapons":       read_sheet(g("Weapons"))       if g("Weapons")       else [],
  "weapon_levels": read_sheet(g("WeaponLevels"))  if g("WeaponLevels")  else [],
  "items":         read_sheet(g("Items"))         if g("Items")         else [],
  "arcs":          read_sheet(g("Arcs"))          if g("Arcs")          else [],
  "arc_drops":     read_sheet(g("ArcDrops"))      if g("ArcDrops")      else [],
  "meta": {"source_excel": a.excel}
}

out = Path(a.out); out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out}")

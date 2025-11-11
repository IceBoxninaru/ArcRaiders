from pathlib import Path
from openpyxl import load_workbook
import csv, sys

excel = Path("arcraiders_workbook_v2.xlsx")
if not excel.exists():
    print("skip: no excel"); sys.exit(0)
wb = load_workbook(excel, data_only=True)

def dump(sheet, out):
    if sheet not in wb.sheetnames:
        Path(out).write_text("", encoding="utf-8"); return
    ws = wb[sheet]
    heads = [c.value if c.value is not None else "" for c in ws[1]]
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(heads)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r): continue
            w.writerow([("" if v is None else v) for v in r])

dump("Weapons",      "data/weapons.csv")
dump("WeaponLevels", "data/weaponlevels.csv")
dump("Items",        "data/items.csv")
dump("Arcs",         "data/arcs.csv")
dump("ArcDrops",     "data/arcdrops.csv")
print("exported CSVs into data/")

from pathlib import Path
from openpyxl import load_workbook, Workbook

excel = Path("arcraiders_workbook_v2.xlsx")
if excel.exists():
    wb = load_workbook(excel)
else:
    wb = Workbook()
    wb.remove(wb.active)

def ensure_sheet(name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(headers)
        return ws
    ws = wb.create_sheet(name)
    ws.append(headers); return ws

items = ensure_sheet("Items", ["id","name_ja","name_en","rarity","price","type","description","image_rel","locations"])
hdr = [c.value for c in items[1]]
col = {h:i+1 for i,h in enumerate(hdr)}
have = set()
for r in items.iter_rows(min_row=2, values_only=True):
    if r and r[col["image_rel"]-1]:
        have.add(r[col["image_rel"]-1])

# assets から拾う
assets = Path("site/assets/items")
rows = 0
for p in assets.glob("*"):
    if not p.is_file(): continue
    if p.suffix.lower() not in {".png",".jpg",".jpeg",".webp",".gif",".avif",".svg"}: continue
    rel = f"items/{p.name}"
    if rel in have: continue
    stem = p.stem
    items.append([stem,"","","","", "", "", rel, ""])
    rows += 1

wb.save(excel)
print(f"seeded {rows} rows into Items")

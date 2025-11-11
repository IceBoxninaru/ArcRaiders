#!/usr/bin/env python3
import csv, json
from pathlib import Path

CSV_DIR = Path("data")
EXCEL = Path("arcraiders_workbook_v2.xlsx")
OUT = Path("site/data/arcraiders.json")
SHEETS = ["Weapons","WeaponLevels","Items","Arcs","ArcDrops"]

def read_csv(name):
    p = CSV_DIR / f"{name.lower()}.csv"
    if not p.exists(): return []
    rows=[]
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if any((v or "").strip() for v in r.values()):
                rows.append(r)
    return rows

def read_excel():
    from openpyxl import load_workbook
    if not EXCEL.exists():
        return {k:[] for k in ["weapons","weapon_levels","items","arcs","arc_drops"]}
    wb = load_workbook(EXCEL, data_only=True)
    def grab(n):
        if n not in wb.sheetnames: return []
        ws = wb[n]; heads=[c.value for c in ws[1]]; out=[]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r): continue
            out.append({h:v for h,v in zip(heads,r)})
        return out
    return {
        "weapons": grab("Weapons"),
        "weapon_levels": grab("WeaponLevels"),
        "items": grab("Items"),
        "arcs": grab("Arcs"),
        "arc_drops": grab("ArcDrops"),
    }

def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    csv_present = any((CSV_DIR / f"{n.lower()}.csv").exists() for n in SHEETS)
    if csv_present:
        data = {
            "weapons": read_csv("Weapons"),
            "weapon_levels": read_csv("WeaponLevels"),
            "items": read_csv("Items"),
            "arcs": read_csv("Arcs"),
            "arc_drops": read_csv("ArcDrops"),
        }
    else:
        data = read_excel()
    data["meta"] = {"source": "csv" if csv_present else "excel"}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")
if __name__ == "__main__": main()
